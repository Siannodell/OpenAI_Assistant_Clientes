import openai
import streamlit as st
from bs4 import BeautifulSoup
from streamlit.components.v1 import html

import requests
import pdfkit
import time
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


load_dotenv()
#id do assistente
assistant_id = "asst_PeSYlhe0cOtt8YgGczwhnKJ2"

# inicializa cliente openai
client = openai

# inicializa a sessão para ler os ids
if "file_id_list" not in st.session_state:
    st.session_state.file_id_list = []

if "start_chat" not in st.session_state:
    st.session_state.start_chat = False

if "thread_id" not in st.session_state:
    st.session_state.thread_id = None

# titulo e icone da página
# Função para converter XLSX pra PDF
def convert_xlsx_to_pdf(input_path, output_path):
    workbook = load_workbook(input_path)
    sheets = workbook.sheetnames

    pdf = canvas.Canvas(output_path, pagesize=letter)

    for sheet_name in sheets:
        sheet = workbook[sheet_name]

        for row in sheet.iter_rows():
            for cell in row:
                pdf.drawString(cell.column * 50, letter[1] - cell.row * 10, str(cell.value))

    pdf.save()

# Função pra enviar arquivo convertido pra OpenAI
def upload_to_openai(filepath):
    with open(filepath, "rb") as file:
        response = openai.files.create(file=file.read(), purpose="assistants")
    return response.id

#local
#api_key = os.getenv("OPENAI_API_KEY")
#git
api_key = st.secrets.OpenAIAPI.openai_api_key
if api_key:
    openai.api_key = api_key

# Execute your app
icon_copy = """
<svg xmlns="http://www.w3.org/2000/svg" xmlns:xlink="http://www.w3.org/1999/xlink" fill="white" height="16px" width="16px" version="1.1" viewBox="0 0 512 512" enable-background="new 0 0 512 512">
  <g>
    <g>
      <path d="M480.6,109.1h-87.5V31.4c0-11.3-9.1-20.4-20.4-20.4H31.4C20.1,11,11,20.1,11,31.4v351c0,11.3,9.1,20.4,20.4,20.4h87.5    v77.7c0,11.3,9.1,20.4,20.4,20.4h341.3c11.3,0,20.4-9.1,20.4-20.4v-351C501,118.3,491.9,109.1,480.6,109.1z M51.8,362V51.8h300.4    v57.3H139.3c-11.3,0-20.4,9.1-20.4,20.4V362H51.8z M460.2,460.2H159.7V150h300.4V460.2z"/>
      <path d="m233.3,254.4h155.8c11.3,0 20.4-9.1 20.4-20.4 0-11.3-9.1-20.4-20.4-20.4h-155.8c-11.3,0-20.4,9.1-20.4,20.4 0,11.2 9.1,20.4 20.4,20.4z"/>
      <path d="m233.3,396.6h155.8c11.3,0 20.4-9.1 20.4-20.4 0-11.3-9.1-20.4-20.4-20.4h-155.8c-11.3,0-20.4,9.1-20.4,20.4 0,11.3 9.1,20.4 20.4,20.4z"/>
    </g>
  </g>
</svg>
"""

perguntas = [
    "Sugestão de pergunta 1",
    "Sugestão de pergunta 2",
    "Sugestão de pergunta 3",
]

st.sidebar.write("<a style='color:white'  href='https://www.google.com.br/' id='baixarArquivo'>[Baixe o arquivo para fazer a análise]</a>", unsafe_allow_html=True)

uploaded_file = st.sidebar.file_uploader("Envie um arquivo", key="file_uploader")

if st.sidebar.button("Enviar arquivo"):
    if uploaded_file:
        # Converter XLSX para PDF
        pdf_output_path = "converted_file.pdf"
        convert_xlsx_to_pdf(uploaded_file, pdf_output_path)

        # Enviar o arquivo convertido
        additional_file_id = upload_to_openai(pdf_output_path)
        
        st.session_state.file_id_list.append(additional_file_id)
        st.sidebar.write(f"ID do arquivo: {additional_file_id}")
        
# Mostra os ids
if st.session_state.file_id_list:
    st.sidebar.write("IDs dos arquivos enviados:")
    for file_id in st.session_state.file_id_list:
        st.sidebar.write(file_id)
        # Associa os arquivos ao assistente
        assistant_file = client.beta.assistants.files.create(
            assistant_id=assistant_id, 
            file_id=file_id
        )

# Botão para iniciar o chat
if st.sidebar.button("Iniciar chat"):
    # Verifica se o arquivo foi upado antes de iniciar
    if st.session_state.file_id_list:
        st.session_state.start_chat = True
        # Cria a thread e guarda o id na sessão
        thread = client.beta.threads.create()
        st.session_state.thread_id = thread.id
        st.write("id da thread: ", thread.id)
    else:
        st.sidebar.warning("Por favor, selecione pelo menos um arquivo para iniciar o chat")


if st.session_state.start_chat:
    on = st.sidebar.toggle('Ver sugestões de perguntas')

    if on:
        for indice, pergunta in enumerate(perguntas):
            st.sidebar.write(f"<a style=\"color:white;display:flex;align-items:center;gap:26px;text-decoration:none\" target=\"_self\" id=\"pergunta{indice}\" href=\"javascript:(function(){{var conteudo = document.getElementById('pergunta{indice}').innerText; navigator.clipboard.writeText(conteudo).then(function() {{ console.log('Conteúdo copiado para a área de transferência: ' + conteudo); }}, function(err) {{ console.error('Erro ao copiar conteúdo: ', err); }});}})()\">{pergunta}<span>{icon_copy}</span></a>", unsafe_allow_html=True)

    st.sidebar.write('<style>.st-bx {background: #282828}</style>', unsafe_allow_html=True)

# Função para copiar o texto para a área de transferência
def copy_to_clipboard(text):
    # Copiar o texto para a área de transferência
    st.write(text, key="text_to_copy")

# Obtendo a pergunta e o ícone
pergunta = "Sua pergunta aqui"
icon_copy = "📋"

# Botão para copiar o texto
if st.button(f"Copy: {pergunta} {icon_copy}"):
    copy_to_clipboard(pergunta)

# Exibindo a pergunta
st.write(pergunta)

# Define a função para iniciar
def process_message_with_citations(message):
    """Extract content and annotations from the message and format citations as footnotes."""
    message_content = message.content[0].text
    annotations = message_content.annotations if hasattr(message_content, 'annotations') else []
    citations = []

    # for nas annotations
    for index, annotation in enumerate(annotations):
        # substitui o texto da mensagem
        message_content.value = message_content.value.replace(annotation.text, f' [{index + 1}]')

        if (file_citation := getattr(annotation, 'file_citation', None)):
            cited_file = {'filename': 'cited_document.pdf'}  # Substituído pelo arquivo retornado
            citations.append(f'[{index + 1}] {file_citation.quote} from {cited_file["filename"]}')
        elif (file_path := getattr(annotation, 'file_path', None)):
            # Placeholder for file download citation
            cited_file = {'filename': 'downloaded_document.pdf'}  # Substituído pelo arquivo retornado
            citations.append(f'[{index + 1}] Click [here](#) to download {cited_file["filename"]}')  # Link de download substituído pelo caminho do arquivo

    # Adiciona notas no final da mensgaem (talvez tirar)
    full_response = message_content.value
    return full_response

# Interface do chat
st.subheader("ANÁLISE DE CLIENTES")
#st.write("Este chat usa a API da OpenAI para gerar respostas.")

# Só vai mostrar o chat se for iniciado
if st.session_state.start_chat:
    # Inicializa o modelo usado
    if "openai_model" not in st.session_state:
        st.session_state.openai_model = "gpt-4-1106-preview"
    if "messages" not in st.session_state:
        st.session_state.messages = []

    # Mostra mensagens anteriores
    for message in st.session_state.messages:
        with st.chat_message(message["role"]):
            st.markdown(message["content"])

    # Campo pro usuário escrever
    if prompt := st.chat_input("Faça uma pergunta!"):
        # Adiciona as mensagens do usuário e mostra no chat
        st.session_state.messages.append({"role": "user", "content": prompt})
        with st.chat_message("user"):
            st.markdown(prompt)

        # Adiciona as mensagens criadas na thread
        client.beta.threads.messages.create(
            thread_id=st.session_state.thread_id,
            role="user",
            content=prompt
        )

        # Cria a requisição com mais instruções
        run = client.beta.threads.runs.create(
            thread_id=st.session_state.thread_id,
            assistant_id=assistant_id,
            instructions="Por favor, responda as perguntas usando o conteúdo do arquivo. Quando adicionar informações externas, seja claro e mostre essas informações em outra cor."
        )

        # Pedido para finalizar a requisição e retornar as mensagens do assistente
        while run.status != 'completed':
            time.sleep(1)
            run = client.beta.threads.runs.retrieve(
                thread_id=st.session_state.thread_id,
                run_id=run.id
            )

        # Retorna as mensagens do assistente
        messages = client.beta.threads.messages.list(
            thread_id=st.session_state.thread_id
        )

        # Processa e mostra as mensagens do assistente
        assistant_messages_for_run = [
            message for message in messages 
            if message.run_id == run.id and message.role == "assistant"
        ]
        for message in assistant_messages_for_run:
            full_response = process_message_with_citations(message)
            st.session_state.messages.append({"role": "assistant", "content": full_response})
            with st.chat_message("assistant"):
                st.markdown(full_response, unsafe_allow_html=True)
else:
    # Prompt pra iniciar o chat
    st.write("Por favor, selecione o(s) arquivo(s) e clique em *iniciar chat* para gerar respostas")
